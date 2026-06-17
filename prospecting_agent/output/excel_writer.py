import os
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config
from models.lead import Lead
from utils.logger import get_logger

log = get_logger("excel_writer")

HEADER_BLUE = "2E75B6"
ROW_BLUE    = "EBF3FB"

COLUMNS = [
    ("Vertical",              22),
    ("SFDC Account",          22),
    ("Prospect Business Name", 35),
    ("Prospect Link",         40),
]


def _apply_header(ws) -> None:
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(fill_type="solid", fgColor=HEADER_BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center")
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 20


def _write_lead_row(ws, row_idx: int, lead: Lead) -> None:
    fill = PatternFill(fill_type="solid", fgColor=ROW_BLUE) if row_idx % 2 == 0 else None
    normal_font = Font(name="Calibri", size=11)
    link_font   = Font(name="Calibri", size=11, color="0563C1", underline="single")

    def w(col, value, font=None):
        c = ws.cell(row=row_idx, column=col, value=value)
        c.font      = font or normal_font
        c.alignment = Alignment(vertical="center")
        if fill:
            c.fill = fill
        return c

    w(1, lead.industry)   # Vertical
    w(2, "")              # SFDC Account — left blank for the rep to fill in
    w(3, lead.company_name)

    # Prospect Link — clickable hyperlink
    url = lead.website or ""
    c = ws.cell(row=row_idx, column=4, value=url)
    c.alignment = Alignment(vertical="center")
    if fill:
        c.fill = fill
    if url:
        c.hyperlink = url
        c.font = link_font
    else:
        c.font = normal_font

    ws.row_dimensions[row_idx].height = 18


def write_excel(leads: List[Lead], sector_names: List[str]) -> str:
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    now = datetime.now()
    sector_code = "_".join(s[:3].upper() for s in sector_names)
    filename = f"purolator_leads_{sector_code}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(config.REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    _apply_header(ws)

    for row_idx, lead in enumerate(leads, start=2):
        _write_lead_row(ws, row_idx, lead)

    wb.save(filepath)
    log.info("Excel saved: %s (%d leads)", filepath, len(leads))
    return filepath
